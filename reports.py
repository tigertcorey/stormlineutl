"""
Stormline report generator — produces Excel workbooks from estimate data.
Mirrors the PlanSwift Quote / Quote Summary factory reports.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               numbers)
from openpyxl.utils import get_column_letter

NAVY   = "1F3864"
BLUE   = "2E74B5"
LTBLUE = "D9E2F3"
GRAY   = "F2F2F2"
WHITE  = "FFFFFF"
GREEN  = "375623"

_thin = Side(style="thin")
_med  = Side(style="medium")
_border_thin = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_border_top  = Border(top=_med, bottom=_thin, left=_thin, right=_thin)


def _hdr(ws, row, col, text, bold=True, bg=NAVY, fg=WHITE, size=10, wrap=False, align="center"):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Calibri", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _border_top
    return cell


def _row(ws, row, col, value, bold=False, bg=WHITE, fmt=None, align="left", indent=0):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", bold=bold, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", indent=indent)
    cell.border = _border_thin
    if fmt:
        cell.number_format = fmt
    return cell


def generate_quote_report(estimate_data: dict, out_dir: str) -> str:
    """
    Generate Excel quote report (two sheets: Quote Detail + Quote Summary).
    Returns the output file path.
    """
    d = estimate_data.get("data", estimate_data)
    job_name   = d.get("job_name", "Project")
    gc_name    = d.get("gc_name", "")
    line_items = d.get("line_items", [])
    mob        = d.get("mobilization", 0)
    testing    = d.get("testing", 0)
    sales_tax  = d.get("sales_tax", 0)
    total_bid  = d.get("total_bid", 0)
    direct     = d.get("direct_cost", 0)
    oap_pct    = d.get("oap_rate_pct", 20)
    oap_amt    = d.get("oap_amount", 0)
    today      = datetime.now().strftime("%B %d, %Y")

    wb = Workbook()

    # ── Sheet 1: Quote Detail ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Quote Detail"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 7
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16

    # Title block
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "STORMLINE UTILITIES, LLC — QUOTE DETAIL"
    c.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = f"{job_name}  |  GC: {gc_name}  |  Date: {today}"
    c.font = Font(name="Calibri", size=10, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Column headers
    ws.row_dimensions[3].height = 20
    for col, (label, al) in enumerate([
        ("#", "center"), ("Description", "left"), ("Unit", "center"),
        ("Quantity", "center"), ("Unit Price", "right"), ("Total", "right")
    ], 1):
        _hdr(ws, 3, col, label, align=al)

    # Group items by section
    from itertools import groupby
    row = 4
    for section, items in groupby(line_items, key=lambda x: x["section"]):
        # Section header
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value=section.upper())
        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = _border_thin
        ws.row_dimensions[row].height = 16
        row += 1

        sec_items = list(items)
        sec_total = 0.0
        for i, li in enumerate(sec_items, 1):
            bg = GRAY if i % 2 == 0 else WHITE
            qty  = li["qty"]
            uc   = li["unit_cost"]
            ext  = li["extension"]
            sec_total += ext
            _row(ws, row, 1, i,              align="center", bg=bg)
            _row(ws, row, 2, li["item"],     align="left",   bg=bg, indent=1)
            _row(ws, row, 3, li["unit"],     align="center", bg=bg)
            _row(ws, row, 4, qty,            align="right",  bg=bg,
                 fmt="#,##0" if qty == int(qty) else "#,##0.0")
            _row(ws, row, 5, uc,             align="right",  bg=bg, fmt='"$"#,##0.00')
            _row(ws, row, 6, ext,            align="right",  bg=bg, fmt='"$"#,##0.00')
            ws.row_dimensions[row].height = 15
            row += 1

        # Section subtotal
        c1 = ws.cell(row=row, column=1, value="")
        ws.merge_cells(f"A{row}:E{row}")
        c2 = ws.cell(row=row, column=1)
        c2.value = f"{section.upper()} SUBTOTAL"
        c2.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c2.fill = PatternFill("solid", fgColor=NAVY)
        c2.alignment = Alignment(horizontal="right", vertical="center")
        c2.border = _border_thin
        c3 = ws.cell(row=row, column=6, value=sec_total)
        c3.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c3.fill = PatternFill("solid", fgColor=NAVY)
        c3.number_format = '"$"#,##0.00'
        c3.alignment = Alignment(horizontal="right")
        c3.border = _border_thin
        ws.row_dimensions[row].height = 16
        row += 2

    # General conditions rows
    for label, val in [("Mobilization", mob), ("Testing / Chlorination TV", testing), ("Sales Tax (8.25% TX materials)", sales_tax)]:
        _row(ws, row, 1, "", bg=GRAY)
        ws.merge_cells(f"A{row}:E{row}")
        c = ws.cell(row=row, column=1)
        c.value = label
        c.font = Font(name="Calibri", italic=True, size=10)
        c.fill = PatternFill("solid", fgColor=LTBLUE)
        c.alignment = Alignment(horizontal="right", indent=2)
        c.border = _border_thin
        cv = ws.cell(row=row, column=6, value=val)
        cv.font = Font(name="Calibri", italic=True, size=10)
        cv.fill = PatternFill("solid", fgColor=LTBLUE)
        cv.number_format = '"$"#,##0.00'
        cv.alignment = Alignment(horizontal="right")
        cv.border = _border_thin
        ws.row_dimensions[row].height = 15
        row += 1

    row += 1
    # O&P row
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1)
    c.value = f"Overhead & Profit ({oap_pct}%)"
    c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = Alignment(horizontal="right", indent=2)
    c.border = _border_thin
    cv = ws.cell(row=row, column=6, value=oap_amt)
    cv.font = Font(name="Calibri", bold=True, color=WHITE)
    cv.fill = PatternFill("solid", fgColor=BLUE)
    cv.number_format = '"$"#,##0.00'
    cv.alignment = Alignment(horizontal="right")
    cv.border = _border_thin
    ws.row_dimensions[row].height = 16
    row += 1

    # Total row
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1)
    c.value = "TOTAL BASE BID"
    c.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="right", indent=2)
    c.border = _border_thin
    cv = ws.cell(row=row, column=6, value=total_bid)
    cv.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
    cv.fill = PatternFill("solid", fgColor=NAVY)
    cv.number_format = '"$"#,##0.00'
    cv.alignment = Alignment(horizontal="right")
    cv.border = _border_thin
    ws.row_dimensions[row].height = 22

    # ── Sheet 2: Quote Summary ────────────────────────────────────────────
    ws2 = wb.create_sheet("Quote Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 18

    ws2.merge_cells("A1:B1")
    c = ws2["A1"]
    c.value = "STORMLINE UTILITIES — QUOTE SUMMARY"
    c.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    ws2.merge_cells("A2:B2")
    c = ws2["A2"]
    c.value = f"{job_name}  |  {today}"
    c.font = Font(name="Calibri", size=10, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = Alignment(horizontal="center")
    ws2.row_dimensions[2].height = 16

    r = 4
    sec_totals = d.get("section_totals", {})
    for sec, amt in sec_totals.items():
        c1 = ws2.cell(row=r, column=1, value=sec)
        c1.font = Font(name="Calibri", size=11, bold=True)
        c1.fill = PatternFill("solid", fgColor=LTBLUE)
        c1.alignment = Alignment(horizontal="left", indent=1)
        c1.border = _border_thin
        c2 = ws2.cell(row=r, column=2, value=amt)
        c2.font = Font(name="Calibri", size=11, bold=True)
        c2.fill = PatternFill("solid", fgColor=LTBLUE)
        c2.number_format = '"$"#,##0.00'
        c2.alignment = Alignment(horizontal="right")
        c2.border = _border_thin
        ws2.row_dimensions[r].height = 18
        r += 1

    r += 1
    for label, val in [("Mobilization", mob), ("Testing", testing), ("O&P (%d%%)" % oap_pct, oap_amt), ("Sales Tax", sales_tax)]:
        c1 = ws2.cell(row=r, column=1, value=label)
        c1.font = Font(name="Calibri", size=10, italic=True)
        c1.fill = PatternFill("solid", fgColor=GRAY)
        c1.alignment = Alignment(horizontal="left", indent=2)
        c1.border = _border_thin
        c2 = ws2.cell(row=r, column=2, value=val)
        c2.font = Font(name="Calibri", size=10, italic=True)
        c2.fill = PatternFill("solid", fgColor=GRAY)
        c2.number_format = '"$"#,##0.00'
        c2.alignment = Alignment(horizontal="right")
        c2.border = _border_thin
        ws2.row_dimensions[r].height = 15
        r += 1

    r += 1
    c1 = ws2.cell(row=r, column=1, value="TOTAL BASE BID")
    c1.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    c1.fill = PatternFill("solid", fgColor=NAVY)
    c1.alignment = Alignment(horizontal="left", indent=1)
    c1.border = _border_thin
    c2 = ws2.cell(row=r, column=2, value=total_bid)
    c2.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    c2.fill = PatternFill("solid", fgColor=NAVY)
    c2.number_format = '"$"#,##0.00'
    c2.alignment = Alignment(horizontal="right")
    c2.border = _border_thin
    ws2.row_dimensions[r].height = 22

    # Save
    safe = job_name.strip().replace(" ", "_")
    out_path = os.path.join(out_dir, f"QUOTE_{safe}_{datetime.now().strftime('%Y%m%d')}.xlsx")
    os.makedirs(out_dir, exist_ok=True)
    wb.save(out_path)
    return out_path
